import pandas as pd
from dotenv import load_dotenv
import pyodbc
import os
import ast
from openpyxl import load_workbook

from db import create_sql_connection

load_dotenv()

def transform_date(date_value:str):
    date_value_list = date_value.split(", ")
    month_mapping = {month: index for index, month in enumerate(['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'], start=1)}
    if len(date_value_list)>0:
        only_date = date_value_list[1].split(" ")
        day, month_text, year = only_date
        month_number = month_mapping.get(month_text, None)
        if month_number is None:
            raise ValueError(f"Mes '{month_text}' no reconocido")
    return f"{year}-{month_number:02d}-{day.zfill(2)}"

# def csv_to_sql_server(dir_, dir_url, conn_):
#     all_matchs_data = pd.DataFrame()
#     for file_ in dir_:
#         print({'file_url': os.path.join(dir_url, file_)})
#         df = pd.read_csv(os.path.join(dir_url, file_))
#         df['date'] = df['date'].apply(transform_date)
#         print(df.columns)
#         all_matchs_data = pd.concat([all_matchs_data, df], ignore_index=True)
#     table = 'matchs_data'
#     schema = "torneos_futbol"
#     print({'table_columns': all_matchs_data.columns.to_list(), 'table_count': all_matchs_data.shape[0]})
#     all_matchs_data.to_sql(table, conn_, if_exists='replace', schema=schema, index=False)
#     print(f"Table {table} was saved successfully")

def sheet_exists(file_path, sheet_name):
    try:
        workbook = load_workbook(file_path)
        return sheet_name in workbook.sheetnames
    except FileNotFoundError:
        return False
    
def create_players_table(df_: pd.DataFrame, player_col:str, team_col:str):
    players_list = []
    for index, row in df_.iterrows():
        players = row[player_col]
        team = row[team_col]
        for player in players:
            players_list.append({"name": player, "team_ID": team})
    return players_list

def transform_data(dir_, dir_url):
    all_matchs_data = pd.DataFrame()
    for file_ in dir_:
        if any(year in file_ for year in ['2020', '2021', '2022', '2023', '2024', '2025']):
            print({'file_url': os.path.join(dir_url, file_)})
            df = pd.read_csv(os.path.join(dir_url, file_))
            df['date'] = df['date'].apply(transform_date)
            all_matchs_data = pd.concat([all_matchs_data, df], ignore_index=True)
    table = 'matchs_data'
    print({'table_columns': all_matchs_data.columns.to_list(), 'table_count': all_matchs_data.shape[0]})
    all_matchs_data = all_matchs_data[all_matchs_data.status == "FINALIZADO"]
    all_matchs_data.to_csv(f"./CSV/{table}.csv", sep=",", index=False, encoding="utf-8")
    all_matchs_data_filtered = all_matchs_data[(all_matchs_data.tournament.isin(["copa argentina", "copa libertadores", "copa sudamericana", "primera division argentina", "torneos verano argentina", "torneo inicial"])) & (all_matchs_data.status == "FINALIZADO")].reset_index(drop=True)
    all_matchs_data_filtered["id"] = all_matchs_data_filtered.index + 1
    event_types = {
        "local_scorers": "goal_local",
        "away_scorers": "goal_away",
        "local_yellow_cards": "yellow_card_local",
        "away_yellow_cards": "yellow_card_away",
        "local_red_cards": "red_card_local",
        "away_red_cards": "red_card_away"
    }
    event_list = []

    columns_to_split = [
        'local_scorers', 'away_scorers', 
        'local_yellow_cards', 'away_yellow_cards', 
        'local_red_cards', 'away_red_cards'
    ]

    for col in columns_to_split:
        all_matchs_data_filtered.loc[:, col] = all_matchs_data_filtered[col].fillna('').apply(lambda x: str(x).split(', ') if x else [])

    for idx, row in all_matchs_data_filtered.iterrows():
        match_id = row["id"]
        for event_column, event_name in event_types.items():
            players = row[event_column]  # Lista de jugadores directamente
            if isinstance(players, list):  # Verificar que sea una lista
                for player in players:
                    event_list.append({
                        "type": event_name,
                        "match_ID": match_id,
                        "team_ID": row["home_team"] if "local" in event_column else row["away_team"],
                        "player_ID": player
                    })
                    
    matchs_table = all_matchs_data_filtered[[
        'tournament', 'date', 'year', 'home_team', 'away_team', 'score', 'local_ball_position', 'away_ball_position', 'local_goals', 'away_goals', 'local_kicks_to_goals', 'away_kicks_to_goals', 'local_outside_kicks', 'away_outside_kicks', 'local_total_kicks', 'away_total_kicks', 'local_shortcuts', 'away_shortcuts', 'local_corner_kicks', 'away_corner_kicks', 'local_offside', 'away_offside', 'local_red_cards.1', 'away_red_cards.1', 'local_substitutions', 'away_substitutions', 'local_faults', 'away_faults', 'local_assists', 'away_assists', 'local_crossbar_kicks', 'away_crossbar_kicks', 'local_lesions', 'away_lesions', 'local_commited_penalties', 'away_commited_penalties'
    ]].reset_index(drop=True).rename(
        columns={
            "tournament": 'tournament_ID', 
            "home_team": 'home_team_ID', 
            "away_team": 'away_team_ID',
            'local_ball_position': 'home_ball_position',
            'local_goals': 'home_goals',
            'local_kicks_to_goals': 'home_kicks_to_goals',
            'local_outside_kicks': 'home_outside_kicks',
            'local_total_kicks': 'home_total_kicks',
            'local_shortcuts': 'home_shortcuts',
            'local_corner_kicks': 'home_corner_kicks',
            'local_offside': 'home_offside',
            'local_substitutions': 'home_substitutions',
            'local_faults': 'home_faults',
            'local_assists': 'home_assists',
            'local_crossbar_kicks': 'home_crossbar_kicks',
            'local_lesions': 'home_lesions',
            'local_commited_penalties': 'home_commited_penalties'}
    ).drop(columns=[
        'local_red_cards.1', 'away_red_cards.1'
    ], axis=1)
    matchs_table.index = matchs_table.index + 1
    table_2 = 'matchs_data_arg'
    table_3 = 'matchs_data_arg_3'
    all_matchs_data_filtered.to_csv(f"./CSV/{table_2}.csv", sep=",", index=False, encoding="utf-8")
    xlsx_path = f"./XLSX/{table}.xlsx"
    # print(create_players_table(all_matchs_data_filtered))
    players_data = []
    players_data.extend(create_players_table(all_matchs_data_filtered, "local_scorers", "home_team"))
    players_data.extend(create_players_table(all_matchs_data_filtered, "away_scorers", "away_team"))
    players_data.extend(create_players_table(all_matchs_data_filtered, "local_yellow_cards", "home_team"))
    players_data.extend(create_players_table(all_matchs_data_filtered, "away_yellow_cards", "away_team"))
    players_data.extend(create_players_table(all_matchs_data_filtered, "local_red_cards", "home_team"))
    players_data.extend(create_players_table(all_matchs_data_filtered, "away_red_cards", "away_team"))
    df_players = pd.DataFrame(players_data).drop_duplicates(subset='name')
    df_players.index = df_players.index + 1
    df_teams = pd.DataFrame({'name': all_matchs_data_filtered.home_team.unique()})
    df_teams.index = df_teams.index + 1
    df_tournaments = pd.DataFrame({'name': all_matchs_data_filtered.tournament.unique()})
    df_tournaments.index = df_tournaments.index + 1
    df_events = pd.DataFrame(event_list)
    df_events.index = df_events.index + 1
    with pd.ExcelWriter(xlsx_path, engine='openpyxl', mode='w') as writer:
        all_matchs_data.to_excel(writer, sheet_name="Sheet1", index=False)
        all_matchs_data_filtered.to_excel(writer, sheet_name="Sheet2", index=True)
        matchs_table.to_excel(writer, sheet_name="Matchs", index=True, index_label="id")
        df_players.to_excel(writer, sheet_name="Players", index=True, index_label="id")
        df_teams.to_excel(writer, sheet_name="Teams", index=True, index_label="id")
        df_tournaments.to_excel(writer, sheet_name="Tournaments", index=True, index_label="id")
        df_events.to_excel(writer, sheet_name="Events", index=True, index_label="id")
        print(f"-------------------------------------------------------{xlsx_path} WAS CREATED-----------------------------------------------------")

if __name__ == '__main__':
    db_user = os.getenv('SQL_SERVER_USER')
    db_pass = os.getenv('SQL_SERVER_PASS')
    db_host = os.getenv('SQL_SERVER_HOST')
    db_name = os.getenv('SQL_SERVER_DB')
    database_url = f"mssql+pyodbc://{db_user}:{db_pass}@{db_host}/{db_name}?driver=ODBC+Driver+17+for+SQL+Server"
    engine = create_sql_connection(database_url)
    BASE_DIR = "C:\\Users\\PC\\Documents\\Matias\\data_projects\\torneos_primera_division_arg"
    csv_dir_url = BASE_DIR+"\\CSV"
    csv_dir_files = [file for file in os.listdir(csv_dir_url) if '.csv' in file]
    # curso_data_analytics = 
    # connection_string = f'DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={db_host};DATABASE={db_name};UID={db_user};PWD={db_pass}'
    transform_data(csv_dir_files, csv_dir_url)